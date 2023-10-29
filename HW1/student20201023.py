#!/usr/bin/python3

from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook(filename = 'student.xlsx')
ws = wb.active

count = 74
a_count = int(count * 0.3)
ab_count = int(count * 0.7)
f_count = 0
score = []
score_tmp = []
score_grade = {}

for row in range(2, 76):
    score.append(ws.cell(column = 3, row = row).value * 0.3 +
    ws.cell(column = 4, row = row).value * 0.35 +
    ws.cell(column = 5, row = row).value * 0.34 +
    ws.cell(column = 6, row = row).value)

i = 2
for s in score:
    ws.cell(column = 7, row = i, value = s)
    score_tmp.append(s)
    i += 1

score_tmp.sort(reverse = True)
for s in score_tmp:
    if s < 40:
        f_count += 1

i = 1
for s in score_tmp:
    if i <= a_count:
        if i <= a_count // 2:
            score_grade[s] = 'A+'
        else:
            score_grade[s] = 'A0'
    elif i <= ab_count:
        if i <= (ab_count - a_count) // 2 + a_count:
            score_grade[s] = 'B+'
        else:
            score_grade[s] = 'B0'
    elif i <= count - f_count:
        if i <= ab_count + (count - ab_count - f_count) // 2:
            score_grade[s] = 'C+'
        else:
            score_grade[s] = 'C0'
    else:
        score_grade[s] = 'F'
    i += 1

i = 2
for s in score:
    grade = score_grade.get(s)
    ws.cell(column = 8, row = i, value = grade)
    i += 1

wb.save("student.xlsx")
